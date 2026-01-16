import streamlit as st
import pandas as pd
import re
import io
from openpyxl.styles import Alignment, PatternFill, Border, Side

def extract_area_logic(text):
    if pd.isna(text) or text == "": return 0.0
    
    # 1. Cleanup: Standardize spaces and remove commas in numbers (e.g., 13,600 -> 13600)
    text = " ".join(str(text).split())
    text = re.sub(r'(\d),(\d)', r'\1\2', text) 
    text = text.replace(' ,', ',').replace(', ', ',')
    
    # 2. Refined Focus: Cut off the heavy land/plot/FSI description
    # This jumps straight to the building name or unit details
    focus_keywords = r'(?:इमारतीमधील|अपार्टमेंटमधील|सदनिका|फ्लॅट|युनिट|टावर|टॉवर|flat|unit)'
    parts = re.split(focus_keywords, text, flags=re.IGNORECASE)
    # Use the portion of text starting from the flat/building mention
    relevant_text = " ".join(parts[1:]) if len(parts) > 1 else text

    # Define regex patterns
    m_unit = r'(?:चौ\.?\s*मी\.?|चौरस\s*मी[टत]र|sq\.?\s*m(?:tr)?\.?)'
    f_unit = r'(?:चौ\.?\s*फू\.?|चौरस\s*फु[टत]|sq\.?\s*f(?:t)?\.?)'
    # Adding "सेलेबल" to total keywords
    total_keywords = r'(?:ए[ककु]ण\s*क्षेत्र|क्षेत्रफळ|total\s*area|सेलेबल\s*क्षेत्र)'
    parking_keywords = ["पार्किंग", "पार्कींग", "parking", "पार्कीग", "पार्किंगसह"]

    # --- STRATEGY A: LOOK FOR TOTAL SALEABLE AREA FIRST (Often most accurate in luxury) ---
    # Specifically looking for "एकूण सेलेबल क्षेत्र 13600 चौ. फुट"
    t_f_match = re.search(rf'{total_keywords}\s*(\d+\.?\d*)\s*{f_unit}', relevant_text, re.IGNORECASE)
    if t_f_match:
        return round(float(t_f_match.group(1)) / 10.764, 3)

    t_m_match = re.search(rf'{total_keywords}\s*(\d+\.?\d*)\s*{m_unit}', relevant_text, re.IGNORECASE)
    if t_m_match:
        return round(float(t_m_match.group(1)), 3)

    # --- STRATEGY B: SUM INDIVIDUAL METRIC COMPONENTS ---
    # We find all numbers followed by sq.mt in the relevant section
    m_matches = re.findall(rf'(\d+\.?\d*)\s*{m_unit}', relevant_text, re.IGNORECASE)
    m_vals = []
    
    # Re-iterate to check context for each match
    segments = re.split(f'(\d+\.?\d*)\s*{m_unit}', relevant_text, flags=re.IGNORECASE)
    for i in range(1, len(segments), 2):
        val = float(segments[i])
        context_before = segments[i-1].lower()
        is_parking = any(word in context_before for word in parking_keywords)
        
        # In this duplex case, carpet areas are around 551 and 402. 
        # We increase the threshold to 1000 to catch luxury units but ignore large land.
        if 0 < val < 1000 and not is_parking:
            m_vals.append(val)

    if m_vals:
        # Check if the last value is a sum of previous ones (to avoid double counting)
        if len(m_vals) > 1 and abs(m_vals[-1] - sum(m_vals[:-1])) < 1.0:
            return round(m_vals[-1], 3)
        return round(sum(m_vals), 3)

    # --- STRATEGY C: SUM INDIVIDUAL IMPERIAL COMPONENTS ---
    f_segments = re.split(f'(\d+\.?\d*)\s*{f_unit}', relevant_text, flags=re.IGNORECASE)
    f_vals = []
    for i in range(1, len(f_segments), 2):
        val = float(f_segments[i])
        context_before = f_segments[i-1].lower()
        is_parking = any(word in context_before for word in parking_keywords)
        if 0 < val < 10000 and not is_parking:
            f_vals.append(val)
                
    if f_vals:
        if len(f_vals) > 1 and abs(f_vals[-1] - sum(f_vals[:-1])) < 10:
            return round(f_vals[-1] / 10.764, 3)
        return round(sum(f_vals) / 10.764, 3)
        
    return 0.0

# Remaining functions (determine_config, apply_excel_formatting, Streamlit UI)
# remain the same as the previous version...

def determine_config(area, t1, t2, t3):
    if area == 0: return "N/A"
    if area < t1: return "1 BHK"
    elif area < t2: return "2 BHK"
    elif area < t3: return "3 BHK"
    else: return "4 BHK"

def apply_excel_formatting(df, writer, sheet_name, is_summary=True):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    colors = ["A2D2FF", "FFD6A5", "CAFFBF", "FDFFB6", "FFADAD", "BDB2FF", "9BF6FF"]
    
    color_idx = 0
    start_row_prop = 2
    start_row_cfg = 2
    
    for i in range(1, worksheet.max_row + 1):
        for j in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=i, column=j)
            cell.alignment = center_align
            if is_summary: cell.border = thin_border

    if is_summary:
        for i in range(2, len(df) + 2):
            curr_prop = df.iloc[i-2, 0]
            next_prop = df.iloc[i-1, 0] if i-1 < len(df) else None
            fill = PatternFill(start_color=colors[color_idx % len(colors)], end_color=colors[color_idx % len(colors)], fill_type="solid")
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=i, column=col).fill = fill
            if curr_prop != next_prop:
                if i >= start_row_prop: worksheet.merge_cells(start_row=start_row_prop, start_column=1, end_row=i, end_column=1)
                color_idx += 1
                start_row_prop = i + 1
            
            curr_cfg_key = [df.iloc[i-2, 0], df.iloc[i-2, 1]]
            next_cfg_key = [df.iloc[i-1, 0], df.iloc[i-1, 1]] if i-1 < len(df) else None
            if curr_cfg_key != next_cfg_key:
                if i >= start_row_cfg: worksheet.merge_cells(start_row=start_row_cfg, start_column=2, end_row=i, end_column=2)
                start_row_cfg = i + 1

# --- STREAMLIT UI ---
st.set_page_config(page_title="Real Estate Dashboard", layout="wide")
st.sidebar.header("Calculation Settings")
loading_factor = st.sidebar.number_input("Loading Factor", min_value=1.0, value=1.35, step=0.001, format="%.3f")
t1 = st.sidebar.number_input("1 BHK Threshold (sqft <)", value=600)
t2 = st.sidebar.number_input("2 BHK Threshold (sqft <)", value=850)
t3 = st.sidebar.number_input("3 BHK Threshold (sqft <)", value=1100)

uploaded_file = st.file_uploader("Upload Excel File (.xlsx)", type="xlsx")

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    clean_cols = {c.lower().strip(): c for c in df.columns}
    desc_col, cons_col, prop_col = clean_cols.get('property description'), clean_cols.get('consideration value'), clean_cols.get('property')
    
    if desc_col and cons_col and prop_col:
        with st.spinner('Calculating...'):
            df['Carpet Area (SQ.MT)'] = df[desc_col].apply(extract_area_logic)
            df['Carpet Area (SQ.FT)'] = (df['Carpet Area (SQ.MT)'] * 10.764).round(3)
            df['Saleable Area'] = (df['Carpet Area (SQ.FT)'] * loading_factor).round(3)
            df['APR'] = df.apply(lambda r: round(r[cons_col]/r['Saleable Area'], 3) if r['Saleable Area'] > 0 else 0, axis=1)
            df['Configuration'] = df['Carpet Area (SQ.FT)'].apply(lambda x: determine_config(x, t1, t2, t3))
            
            valid_df = df[df['Carpet Area (SQ.FT)'] > 0].sort_values([prop_col, 'Configuration', 'Carpet Area (SQ.FT)'])
            
            summary = valid_df.groupby([prop_col, 'Configuration', 'Carpet Area (SQ.FT)']).agg(
                Min_APR=('APR', 'min'), Max_APR=('APR', 'max'), Avg_APR=('APR', 'mean'),
                Median_APR=('APR', 'median'),
                Mode_APR=('APR', lambda x: x.mode().iloc[0] if not x.mode().empty else 0),
                Property_Count=(prop_col, 'count')
            ).reset_index()
            
            summary.columns = ['Property', 'Configuration', 'Carpet Area(SQ.FT)', 'Min. APR', 'Max APR', 'Average of APR', 'Median of APR', 'Mode of APR', 'Count of Property']
            summary[['Min. APR', 'Max APR', 'Average of APR', 'Median of APR', 'Mode of APR']] = summary[['Min. APR', 'Max APR', 'Average of APR', 'Median of APR', 'Mode of APR']].round(3)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                apply_excel_formatting(df, writer, 'Raw Data', is_summary=False)
                apply_excel_formatting(summary, writer, 'Summary', is_summary=True)
            
            st.success("Analysis Complete!")
            st.download_button(label="📥 Download Formatted Excel Report", data=output.getvalue(), file_name="Property_Analysis.xlsx")
